# email_manager.py
"""
Módulo para gestionar correos electrónicos usando IMAP
"""

import imaplib
import email
from email.header import decode_header
from datetime import datetime, timezone
import re

# Importar openpyxl al inicio para evitar problemas de importación tardía
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class EmailManager:
    """Clase para gestionar operaciones con correos electrónicos usando IMAP"""
    
    def __init__(self, imap_connection, email_address):
        self.imap_connection = imap_connection
        self.email_address = email_address
    
    def get_emails_in_date_range(self, start_date, end_date, folder='INBOX'):
        """
        Obtiene correos electrónicos en un rango de fechas específico
        
        Args:
            start_date (datetime): Fecha de inicio
            end_date (datetime): Fecha de fin
            folder (str): Carpeta de correo (INBOX, SENT, etc.)
            
        Returns:
            list: Lista de diccionarios con información de los correos
        """
        print(f"Obteniendo correos desde {start_date.strftime('%Y-%m-%d')} hasta {end_date.strftime('%Y-%m-%d')}...")
        
        try:
            # Seleccionar carpeta
            result = self.imap_connection.select(folder)
            if result[0] != 'OK':
                print(f"Error seleccionando carpeta {folder}")
                return []
            
            # Formatear fechas para búsqueda IMAP
            start_date_str = start_date.strftime('%d-%b-%Y')
            end_date_str = end_date.strftime('%d-%b-%Y')
            
            # Buscar correos en el rango de fechas
            search_criteria = f'(SINCE "{start_date_str}" BEFORE "{end_date_str}")'
            print(f"Criterio de búsqueda: {search_criteria}")
            
            result, message_ids = self.imap_connection.search(None, search_criteria)
            
            if result != 'OK':
                print("Error en la búsqueda de correos")
                return []
            
            # Obtener lista de IDs de mensajes
            message_id_list = message_ids[0].split()
            total_emails = len(message_id_list)
            
            print(f"Se encontraron {total_emails} correos en el rango especificado")
            
            if total_emails == 0:
                return []
            
            emails = []
            processed = 0
            
            # Procesar cada correo (limitar a 100 para no sobrecargar)
            max_emails = min(100, total_emails)
            print(f"Procesando los primeros {max_emails} correos...")
            
            for msg_id in message_id_list[-max_emails:]:  # Obtener los más recientes
                try:
                    # Obtener el correo
                    result, msg_data = self.imap_connection.fetch(msg_id, '(RFC822)')
                    
                    if result == 'OK' and msg_data[0]:
                        email_message = email.message_from_bytes(msg_data[0][1])
                        processed_email = self._process_email(email_message)
                        
                        if processed_email:
                            # Verificar que esté en el rango de fechas correcto
                            if self._is_email_in_date_range(processed_email, start_date, end_date):
                                emails.append(processed_email)
                        
                        processed += 1
                        if processed % 10 == 0:
                            print(f"Procesados {processed}/{max_emails} correos...")
                            
                except Exception as e:
                    print(f"Error procesando correo ID {msg_id}: {str(e)}")
                    continue
            
            print(f"Total de correos procesados: {len(emails)}")
            return emails
            
        except Exception as e:
            print(f"Error obteniendo correos: {str(e)}")
            return []
    
    def _process_email(self, email_message):
        """
        Procesa un correo individual y extrae la información necesaria
        
        Args:
            email_message: Objeto email.message.Message
            
        Returns:
            dict: Información procesada del correo
        """
        try:
            # Fecha de recepción
            date_header = email_message.get('Date')
            if date_header:
                try:
                    # Parsear fecha del correo
                    parsed_date = email.utils.parsedate_to_datetime(date_header)
                    if parsed_date.tzinfo is None:
                        # Si no tiene zona horaria, asumir UTC
                        parsed_date = parsed_date.replace(tzinfo=timezone.utc)
                    
                    # Convertir a hora local
                    local_date = parsed_date.astimezone()
                    date_formatted = local_date.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    date_formatted = 'Fecha no disponible'
            else:
                date_formatted = 'Fecha no disponible'
            
            # Asunto del correo
            subject_header = email_message.get('Subject', 'Sin asunto')
            subject = self._decode_header(subject_header)
            
            # Información del remitente
            from_header = email_message.get('From', 'Remitente desconocido')
            sender_email = self._extract_email_address(from_header)
            
            # Extraer dominio del correo del remitente
            domain = self._extract_domain(sender_email)
            
            return {
                'fecha': date_formatted,
                'asunto': subject,
                'remitente_email': sender_email,
                'dominio_remitente': domain,
                'fecha_objeto': parsed_date if 'parsed_date' in locals() else None
            }
            
        except Exception as e:
            print(f"Error procesando correo: {str(e)}")
            return None
    
    def _decode_header(self, header_value):
        """Decodifica headers de correo que pueden estar codificados"""
        if not header_value:
            return ''
        
        try:
            decoded_parts = decode_header(header_value)
            decoded_string = ''
            
            for part, encoding in decoded_parts:
                if isinstance(part, bytes):
                    if encoding:
                        decoded_string += part.decode(encoding)
                    else:
                        decoded_string += part.decode('utf-8', errors='ignore')
                else:
                    decoded_string += str(part)
            
            return decoded_string.strip()
        except:
            return str(header_value)
    
    def _extract_email_address(self, from_header):
        """Extrae la dirección de correo del header From"""
        if not from_header:
            return 'Remitente desconocido'
        
        # Decodificar el header
        decoded_from = self._decode_header(from_header)
        
        # Buscar dirección de correo usando regex
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        match = re.search(email_pattern, decoded_from)
        
        if match:
            return match.group().lower()
        else:
            return decoded_from.strip()
    
    def _extract_domain(self, email_address):
        """
        Extrae el dominio de una dirección de correo electrónico
        
        Args:
            email_address (str): Dirección de correo electrónico
            
        Returns:
            str: Dominio del correo
        """
        if not email_address or '@' not in email_address:
            return 'Dominio desconocido'
        
        try:
            domain = email_address.split('@')[1].lower()
            return domain
        except IndexError:
            return 'Dominio desconocido'
    
    def _is_email_in_date_range(self, processed_email, start_date, end_date):
        """Verifica si un correo está en el rango de fechas especificado"""
        if not processed_email.get('fecha_objeto'):
            return True  # Si no podemos verificar la fecha, incluirlo
        
        email_date = processed_email['fecha_objeto']
        return start_date <= email_date <= end_date
    
    def export_to_excel(self, emails, filename='correos_exportados.xlsx'):
        """
        Exporta la lista de correos a un archivo Excel
        
        Args:
            emails (list): Lista de correos
            filename (str): Nombre del archivo de salida
        """
        try:
            import pandas as pd
            
            if not emails:
                print("No hay correos para exportar")
                return
            
            # Crear DataFrame
            df = pd.DataFrame(emails)
            # Remover la columna de fecha_objeto si existe
            if 'fecha_objeto' in df.columns:
                df = df.drop('fecha_objeto', axis=1)
            
            # Guardar como Excel
            df.to_excel(filename, index=False, engine='openpyxl')
            print(f"Correos exportados a: {filename}")
            
        except ImportError:
            print("Pandas no disponible. Exportando en formato Excel simple...")
            self._export_to_simple_excel(emails, filename)
        except Exception as e:
            print(f"Error exportando a Excel: {str(e)}")
    
    def _export_to_simple_excel(self, emails, filename):
        """
        Exporta correos a Excel sin usar pandas
        
        Args:
            emails (list): Lista de correos
            filename (str): Nombre del archivo
        """
        if not OPENPYXL_AVAILABLE:
            print("❌ Error: openpyxl no está disponible")
            return
            
        try:
            # Crear workbook y hoja
            wb = Workbook()
            ws = wb.active
            ws.title = "Correos"
            
            # Definir encabezados
            headers = ["Fecha", "Fecha Completa", "Asunto", "Remitente", "Dominio", "Carpeta"]
            
            # Escribir encabezados con formato
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Escribir datos
            for row, email_data in enumerate(emails, 2):
                # Extraer solo la fecha sin hora en formato DD/MM/YYYY
                fecha_completa = email_data['fecha']
                try:
                    # Parsear la fecha completa y convertir a DD/MM/YYYY
                    fecha_obj = datetime.strptime(fecha_completa.split(' ')[0], '%Y-%m-%d')
                    fecha_solo = fecha_obj.strftime('%d/%m/%Y')
                except:
                    fecha_solo = 'Fecha inválida'
                
                # Escribir datos en las columnas
                ws.cell(row=row, column=1, value=fecha_solo)
                ws.cell(row=row, column=2, value=email_data['fecha'])
                ws.cell(row=row, column=3, value=email_data['asunto'])
                ws.cell(row=row, column=4, value=email_data['remitente_email'])
                ws.cell(row=row, column=5, value=email_data['dominio_remitente'])
                ws.cell(row=row, column=6, value=email_data.get('carpeta', 'INBOX'))
            
            # Ajustar ancho de columnas automáticamente
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Limitar el ancho máximo para evitar columnas demasiado anchas
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar archivo
            wb.save(filename)
            print(f"Correos exportados a: {filename}")
            
        except Exception as e:
            print(f"Error exportando archivo: {str(e)}")
    
    def get_user_info(self):
        """
        Obtiene información del usuario autenticado
        
        Returns:
            dict: Información del usuario
        """
        return {
            'nombre': 'Usuario IMAP',
            'email': self.email_address
        }
    
    def get_folder_list(self):
        """
        Obtiene la lista de carpetas disponibles
        
        Returns:
            list: Lista de carpetas
        """
        try:
            result, folders = self.imap_connection.list()
            if result == 'OK':
                folder_list = []
                for folder in folders:
                    # Decodificar nombre de carpeta
                    folder_name = folder.decode().split('"')[-2] if '"' in folder.decode() else folder.decode().split()[-1]
                    folder_list.append(folder_name)
                return folder_list
            return ['INBOX']
        except:
            return ['INBOX']