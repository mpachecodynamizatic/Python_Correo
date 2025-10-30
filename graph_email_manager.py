# graph_email_manager.py
"""
Gestor de correos usando Microsoft Graph API (alternativa a IMAP)
"""

import requests
import json
from datetime import datetime, timezone
from urllib.parse import quote

# Importar openpyxl al inicio para evitar problemas de importación tardía
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class GraphEmailManager:
    """Gestor de correos usando Microsoft Graph API"""
    
    def __init__(self, access_token):
        self.access_token = access_token
        self.headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }
        self.base_url = "https://graph.microsoft.com/v1.0"
    
    def get_emails_in_date_range(self, start_date, end_date, folders=['inbox']):
        """
        Obtiene correos en un rango de fechas usando Graph API
        
        Args:
            start_date (datetime): Fecha de inicio
            end_date (datetime): Fecha de fin
            folders (list): Lista de carpetas a buscar (por defecto solo 'inbox')
            
        Returns:
            list: Lista de correos procesados
        """
        print(f"📥 Obteniendo correos desde {start_date.strftime('%Y-%m-%d')} hasta {end_date.strftime('%Y-%m-%d')}...")
        print("🌐 Usando Microsoft Graph API (compatible con cuentas sin IMAP)")
        
        if isinstance(folders, str):
            folders = [folders]
        
        print(f"📂 Carpetas a revisar: {', '.join(folders)}")
        
        all_emails = []
        
        for folder in folders:
            print(f"\n📁 Procesando carpeta: {folder}")
            folder_emails = self._get_emails_from_folder(start_date, end_date, folder)
            all_emails.extend(folder_emails)
        
        print(f"\n📧 Total de correos obtenidos de todas las carpetas: {len(all_emails)}")
        return all_emails
    
    def _get_emails_from_folder(self, start_date, end_date, folder):
        """Obtiene correos de una carpeta específica"""
        try:
            # Convertir fechas a formato ISO UTC
            start_iso = start_date.astimezone(timezone.utc).isoformat()
            end_iso = end_date.astimezone(timezone.utc).isoformat()
            
            # Construir filtro de fecha
            date_filter = f"receivedDateTime ge {start_iso} and receivedDateTime le {end_iso}"
            
            # Buscar carpeta por nombre si no es 'inbox'
            folder_id = folder
            if folder != 'inbox' and not folder.startswith('AAMk'):  # Si no es un ID directo
                folder_id = self._find_folder_by_name(folder)
                if not folder_id:
                    print(f"   ❌ Carpeta '{folder}' no encontrada")
                    return []
            
            # URL de la API
            url = f"{self.base_url}/me/mailFolders/{folder_id}/messages"
            
            # Parámetros
            params = {
                '$filter': date_filter,
                '$select': 'receivedDateTime,subject,from,sender,bodyPreview',
                '$orderby': 'receivedDateTime desc',
                '$top': 100  # Limitar para mejor rendimiento
            }
            
            emails = []
            page_count = 1
            
            while url and page_count <= 5:  # Máximo 5 páginas (500 correos) por carpeta
                print(f"   📄 Procesando página {page_count}...")
                
                response = requests.get(url, headers=self.headers, params=params if page_count == 1 else None)
                
                if response.status_code == 200:
                    data = response.json()
                    page_emails = data.get('value', [])
                    
                    for email_data in page_emails:
                        processed = self._process_email(email_data)
                        if processed:
                            processed['carpeta'] = folder  # Agregar info de carpeta
                            emails.append(processed)
                    
                    print(f"      ✅ {len(page_emails)} correos en esta página")
                    
                    # Siguiente página
                    url = data.get('@odata.nextLink')
                    page_count += 1
                    params = None  # Solo usar params en primera página
                    
                elif response.status_code == 404:
                    print(f"   ❌ Carpeta '{folder}' no encontrada (404)")
                    break
                else:
                    print(f"   ❌ Error API: {response.status_code}")
                    break
            
            print(f"   📧 Total de correos obtenidos de '{folder}': {len(emails)}")
            return emails
            
        except Exception as e:
            print(f"   ❌ Error obteniendo correos de '{folder}': {str(e)}")
            return []
    
    def _process_email(self, email_data):
        """Procesa un correo individual"""
        try:
            # Fecha
            date_str = email_data.get('receivedDateTime')
            if date_str:
                date_obj = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                local_date = date_obj.astimezone()
                formatted_date = local_date.strftime('%Y-%m-%d %H:%M:%S')
            else:
                formatted_date = 'Fecha no disponible'
            
            # Asunto
            subject = email_data.get('subject', 'Sin asunto')
            
            # Remitente
            from_data = email_data.get('from', {})
            email_address_data = from_data.get('emailAddress', {})
            sender_email = email_address_data.get('address', 'Remitente desconocido')
            
            # Dominio
            domain = self._extract_domain(sender_email)
            
            return {
                'fecha': formatted_date,
                'asunto': subject,
                'remitente_email': sender_email,
                'dominio_remitente': domain
            }
            
        except Exception as e:
            print(f"⚠️  Error procesando correo: {str(e)}")
            return None
    
    def _find_folder_by_name(self, folder_name):
        """Busca una carpeta por su nombre y retorna su ID"""
        try:
            response = requests.get(f"{self.base_url}/me/mailFolders", headers=self.headers)
            
            if response.status_code == 200:
                data = response.json()
                
                for folder in data.get('value', []):
                    if folder.get('displayName', '').lower() == folder_name.lower():
                        print(f"   ✅ Carpeta '{folder_name}' encontrada: {folder.get('displayName')}")
                        return folder.get('id')
                
                # Buscar en subcarpetas si no se encuentra en el nivel raíz
                return self._search_subfolders(folder_name, data.get('value', []))
            
            return None
            
        except Exception as e:
            print(f"   ❌ Error buscando carpeta '{folder_name}': {str(e)}")
            return None
    
    def _search_subfolders(self, folder_name, folders):
        """Busca recursivamente en subcarpetas"""
        for folder in folders:
            folder_id = folder.get('id')
            
            try:
                # Obtener subcarpetas
                response = requests.get(f"{self.base_url}/me/mailFolders/{folder_id}/childFolders", headers=self.headers)
                
                if response.status_code == 200:
                    subfolders_data = response.json()
                    
                    for subfolder in subfolders_data.get('value', []):
                        if subfolder.get('displayName', '').lower() == folder_name.lower():
                            print(f"   ✅ Carpeta '{folder_name}' encontrada en subcarpetas: {subfolder.get('displayName')}")
                            return subfolder.get('id')
                    
                    # Buscar recursivamente en subcarpetas
                    result = self._search_subfolders(folder_name, subfolders_data.get('value', []))
                    if result:
                        return result
                        
            except Exception:
                continue
        
        return None
    
    def _extract_domain(self, email_address):
        """Extrae dominio del email"""
        if not email_address or '@' not in email_address:
            return 'Dominio desconocido'
        
        try:
            return email_address.split('@')[1].lower()
        except:
            return 'Dominio desconocido'
    
    def export_to_excel(self, emails, filename=None):
        """Exporta correos a Excel incluyendo información de carpeta"""
        if not emails:
            print("📝 No hay correos para exportar")
            return
        
        # Verificar si openpyxl está disponible
        if not OPENPYXL_AVAILABLE:
            print("❌ Error: openpyxl no está instalado. Instalando...")
            try:
                import subprocess
                subprocess.check_call([
                    "C:/Users/mpacheco/Documents/GITHUB/PYTHON/Python_Correo/.venv/Scripts/python.exe",
                    "-m", "pip", "install", "openpyxl"
                ])
                print("✅ openpyxl instalado correctamente. Reinicia la aplicación.")
                return
            except Exception as e:
                print(f"❌ Error instalando openpyxl: {e}")
                return
        
        # Generar nombre de archivo con timestamp si no se proporciona
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"correos_{timestamp}.xlsx"
        
        try:
            # Crear workbook y hoja
            wb = Workbook()
            ws = wb.active
            ws.title = "Correos"
            
            # Definir encabezados
            headers = ["Fecha", "Fecha Completa", "Asunto", "Remitente", "Dominio", "Carpeta"]
            
            # Escribir encabezados con formato
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Escribir datos
            for row, email_data in enumerate(emails, 2):
                # Extraer solo la fecha sin hora en formato DD/MM/YYYY
                fecha_completa = email_data['fecha']
                try:
                    # Parsear la fecha completa y convertir a DD/MM/YYYY
                    fecha_obj = datetime.strptime(fecha_completa.split(' ')[0], '%Y-%m-%d')
                    fecha_solo = fecha_obj.strftime('%d/%m/%Y')
                except:
                    fecha_solo = 'Fecha inválida'
                
                # Escribir datos en las columnas
                ws.cell(row=row, column=1, value=fecha_solo)
                ws.cell(row=row, column=2, value=email_data['fecha'])
                ws.cell(row=row, column=3, value=email_data['asunto'])
                ws.cell(row=row, column=4, value=email_data['remitente_email'])
                ws.cell(row=row, column=5, value=email_data['dominio_remitente'])
                ws.cell(row=row, column=6, value=email_data.get('carpeta', 'inbox'))
            
            # Ajustar ancho de columnas automáticamente
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Limitar el ancho máximo para evitar columnas demasiado anchas
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar archivo
            wb.save(filename)
            print(f"💾 Correos exportados a: {filename}")
            
        except Exception as e:
            print(f"❌ Error exportando: {str(e)}")
    
    def get_user_info(self):
        """Obtiene info del usuario"""
        try:
            response = requests.get(f"{self.base_url}/me", headers=self.headers)
            
            if response.status_code == 200:
                data = response.json()
                return {
                    'nombre': data.get('displayName', 'Usuario'),
                    'email': data.get('mail') or data.get('userPrincipalName', 'No disponible')
                }
            else:
                return {'nombre': 'Usuario Graph API', 'email': 'No disponible'}
                
        except:
            return {'nombre': 'Usuario Graph API', 'email': 'No disponible'}
    
    def get_folder_list(self):
        """Obtiene lista de carpetas disponibles"""
        try:
            response = requests.get(f"{self.base_url}/me/mailFolders", headers=self.headers)
            
            if response.status_code == 200:
                data = response.json()
                folders = []
                
                for folder in data.get('value', []):
                    folder_info = {
                        'id': folder.get('id'),
                        'name': folder.get('displayName'),
                        'total': folder.get('totalItemCount', 0),
                        'unread': folder.get('unreadItemCount', 0)
                    }
                    folders.append(folder_info)
                
                return folders
            else:
                return [{'id': 'inbox', 'name': 'Bandeja de entrada', 'total': 0, 'unread': 0}]
                
        except:
            return [{'id': 'inbox', 'name': 'Bandeja de entrada', 'total': 0, 'unread': 0}]